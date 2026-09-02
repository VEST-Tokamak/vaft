"""Preset-based tabular summaries and database-independent exports."""

from __future__ import annotations

from vaft.data.eqdsk import TWO_PI, ods_psi_to_wb_per_radian_factor

from dataclasses import dataclass
import json
import logging
import os
from pathlib import Path
import tempfile
from typing import Callable, Sequence
import warnings

import numpy as np
import pandas as pd

import vaft


logger = logging.getLogger(__name__)


EQUILIBRIUM_GLOBAL_COLUMNS = (
    "shot",
    "eq_index",
    "time_s",
    "ip_kA",
    "psi_axis_Wb",
    "q_axis",
    "q_95",
    "q_min",
    "beta_pol",
    "beta_tor",
    "beta_normal",
    "li_3",
    "energy_mhd_J",
    "area_m2",
    "volume_m3",
    "major_radius_m",
    "minor_radius_m",
    "aspect_ratio",
    "elongation",
    "triangularity",
    "triangularity_upper",
    "triangularity_lower",
    "magnetic_axis_r_m",
    "magnetic_axis_z_m",
    "magnetic_axis_btor_T",
    "vacuum_b0_T",
    "vacuum_r0_m",
    "dia_flux_Wb",
    "virial_s1",
    "virial_s2",
    "virial_s3",
    "virial_alpha",
    "virial_b_pa_T",
    "virial_mui",
    "virial_rt_m",
    "virial_phi_dia_comp_Wb",
    "virial_volume_m3",
    "virial_beta_pd_vir",
    "virial_beta",
    "virial_li",
    "virial_beta_lao",
    "virial_li_lao",
    "virial_beta_bongard",
    "virial_li_bongard",
)

# These paths cover both directly exported values and inputs consumed by the
# boundary, stored-energy, diamagnetic-flux, and virial calculations.  The lazy
# ODS backend fetches only leaves actually touched below or by those helpers.
EQUILIBRIUM_GLOBAL_PATHS = (
    "equilibrium.time",
    "equilibrium.time_slice",
    "equilibrium.vacuum_toroidal_field",
)

VACUUM_REFERENCE_RADIUS_M = 0.4
ELEMENTARY_CHARGE_C = 1.602176634e-19

CORE_PROFILES_COLUMNS = (
    "shot",
    "cp_index",
    "eq_index",
    "time_s",
    "ip_kA",
    "b_t_T",
    "p_loss_MW",
    "tau_e_s",
    "tau_e_ipb89_s",
    "tau_e_h98y2_s",
    "tau_e_nstx_s",
    "tau_e_nstx2006l_s",
    "tau_e_kurskiev2022_s",
    "ne_line_1e19_m3",
    "ne_volume_1e19_m3",
    "te_mean_eV",
    "major_radius_m",
    "inverse_aspect_ratio",
    "elongation",
)

CORE_PROFILES_PATHS = (
    "equilibrium",
    "core_profiles",
    "magnetics",
    "tf",
)

VOLUME_AVERAGED_COLUMNS = (
    "shot",
    "cp_index",
    "eq_index",
    "time_core_s",
    "time_equilibrium_s",
    "time_diff_s",
    "ne_volume_m3",
    "te_volume_eV",
    "electron_pressure_volume_Pa",
    "equilibrium_pressure_volume_Pa",
    "electron_to_equilibrium_pressure_ratio",
)

VOLUME_AVERAGED_PATHS = ("equilibrium", "core_profiles")

EFIT_RELIABILITY_COLUMNS = (
    "shot",
    "eq_index",
    "time_s",
    "ip_kA",
    "measurement_index",
    "measurement_type",
    "identifier",
    "constraint_source",
    "measured",
    "reconstructed",
    "residual",
    "absolute_residual",
    "uncertainty",
    "normalized_residual",
    "weight",
    "chi_squared",
    "exact",
    "disabled",
    "chi_squared_reduced",
    "convergence_iterations_n",
    "convergence_deviation",
    "equilibrium_source",
    "equilibrium_lineage",
)

EFIT_MAGNETIC_FAMILIES = (
    ("bpol_probe", True),
    ("flux_loop", True),
    ("pf_current", True),
    ("ip", False),
    ("diamagnetic_flux", False),
)

EFIT_KINETIC_FAMILIES = (
    ("pressure", True),
    ("pressure_rotational", True),
    ("j_tor", True),
    ("mse_polarisation_angle", True),
)

SHOT_OVERVIEW_COLUMNS = (
    "shot",
    "plasma_onset_time_s",
    "pulse_duration_s",
    "max_ip_kA",
    "mean_b_t_T",
)


@dataclass(frozen=True)
class SummaryPreset:
    columns: tuple[str, ...]
    paths: tuple[str, ...]
    key_columns: tuple[str, ...]
    replace_groups: tuple[str, ...]
    sort_columns: tuple[str, ...]
    extractor: Callable[[object, int], list[dict]]


def _safe_get(container, key: str, default=np.nan):
    try:
        return container[key]
    except Exception:
        return default


def _as_float(value) -> float:
    try:
        array = np.asarray(value, dtype=float)
        if array.size == 0:
            return np.nan
        return float(array.reshape(-1)[0])
    except Exception:
        return np.nan


def _extract_time(eq_slice, eq_times, index: int) -> float:
    if "time" in eq_slice:
        return _as_float(eq_slice["time"])
    if eq_times is not None and index < len(eq_times):
        return _as_float(eq_times[index])
    return float(index)


def _extract_q_min(eq_slice) -> float:
    value = _as_float(_safe_get(eq_slice, "global_quantities.q_min", np.nan))
    if np.isfinite(value):
        return value
    return _as_float(_safe_get(eq_slice, "global_quantities.q_min.value", np.nan))


def _virial_values(outputs: dict, index: int) -> dict[str, float]:
    values = outputs.get(index, {})

    def get(*names):
        value = next((values[name] for name in names if name in values), np.nan)
        return _as_float(value)

    return {
        "virial_s1": get("s_1"),
        "virial_s2": get("s_2"),
        "virial_s3": get("s_3"),
        "virial_alpha": get("alpha"),
        "virial_b_pa_T": get("B_pa"),
        "virial_mui": get("mui", "mui_hat"),
        "virial_rt_m": get("rt"),
        "virial_phi_dia_comp_Wb": get("phi_dia_comp"),
        "virial_volume_m3": get("V_p"),
        "virial_beta_pd_vir": get("beta_pd_vir"),
        "virial_beta_lao": get("beta_p_vir_lao", "beta_p_vir"),
        "virial_li_lao": get("li_vir_lao", "li_vir"),
        "virial_beta_bongard": get("beta_p_vir_bongard"),
        "virial_li_bongard": get("li_vir_bongard"),
    }


def _value_at(value, index: int) -> float:
    try:
        array = np.asarray(value, dtype=float).reshape(-1)
        if not array.size:
            return np.nan
        return float(array[index] if index < array.size else array[0])
    except Exception:
        return np.nan


def extract_equilibrium_global(ods, shot: int) -> list[dict]:
    """Extract the legacy equilibrium-global history schema from one lazy ODS."""
    if "equilibrium.time_slice" not in ods or not len(ods["equilibrium.time_slice"]):
        return []

    for updater in (
        vaft.omas.update_equilibrium_boundary,
        vaft.omas.update_equilibrium_global_quantities_q_min,
        vaft.omas.update_equilibrium_global_quantities_volume,
        vaft.omas.update_equilibrium_stored_energy,
    ):
        try:
            updater(ods, time_slice=None)
        except Exception as exc:
            logger.debug("Shot %s: %s failed: %s", shot, updater.__name__, exc)

    try:
        vaft.omas.update_equilibrium_constraints_diamagnetic_flux(ods, time_slice=None)
    except Exception as exc:
        logger.debug("Shot %s: diamagnetic update failed: %s", shot, exc)
    try:
        virial_outputs = vaft.omas.compute_virial_equilibrium_quantities_ods(
            ods, time_slice=None
        )
    except Exception as exc:
        logger.debug("Shot %s: virial calculation failed: %s", shot, exc)
        virial_outputs = {}
    eq_times = (
        np.asarray(ods["equilibrium.time"], dtype=float)
        if "equilibrium.time" in ods
        else None
    )
    b0_source = _safe_get(ods, "equilibrium.vacuum_toroidal_field.b0")
    r0_source = _as_float(_safe_get(ods, "equilibrium.vacuum_toroidal_field.r0"))
    rows: list[dict] = []
    for index in range(len(ods["equilibrium.time_slice"])):
        eq_slice = ods["equilibrium.time_slice"][index]
        major_radius = _as_float(_safe_get(eq_slice, "boundary.geometric_axis.r"))
        minor_radius = _as_float(_safe_get(eq_slice, "boundary.minor_radius"))
        aspect_ratio = (
            major_radius / minor_radius
            if np.isfinite(major_radius)
            and np.isfinite(minor_radius)
            and minor_radius != 0
            else np.nan
        )
        virial = _virial_values(virial_outputs, index)
        source_b0 = _value_at(b0_source, index)
        normalized_b0 = (
            source_b0 * r0_source / VACUUM_REFERENCE_RADIUS_M
            if np.isfinite(source_b0) and np.isfinite(r0_source)
            else np.nan
        )
        row = {
            "shot": int(shot),
            "eq_index": int(index),
            "time_s": _extract_time(eq_slice, eq_times, index),
            "ip_kA": _as_float(_safe_get(eq_slice, "global_quantities.ip")) / 1e3,
            # Column is labeled Wb: convert from the storage convention
            # (Wb for DD-conformant files, Wb/rad for legacy ones; issue #236).
            "psi_axis_Wb": _as_float(_safe_get(eq_slice, "global_quantities.psi_axis"))
            * ods_psi_to_wb_per_radian_factor(eq_slice) * TWO_PI,
            "q_axis": _as_float(_safe_get(eq_slice, "global_quantities.q_axis")),
            "q_95": _as_float(_safe_get(eq_slice, "global_quantities.q_95")),
            "q_min": _extract_q_min(eq_slice),
            "beta_pol": _as_float(_safe_get(eq_slice, "global_quantities.beta_pol")),
            "beta_tor": _as_float(_safe_get(eq_slice, "global_quantities.beta_tor")),
            "beta_normal": _as_float(
                _safe_get(eq_slice, "global_quantities.beta_normal")
            ),
            "li_3": _as_float(_safe_get(eq_slice, "global_quantities.li_3")),
            "energy_mhd_J": _as_float(
                _safe_get(eq_slice, "global_quantities.energy_mhd")
            ),
            "area_m2": _as_float(_safe_get(eq_slice, "global_quantities.area")),
            "volume_m3": _as_float(_safe_get(eq_slice, "global_quantities.volume")),
            "major_radius_m": major_radius,
            "minor_radius_m": minor_radius,
            "aspect_ratio": float(aspect_ratio),
            "elongation": _as_float(_safe_get(eq_slice, "boundary.elongation")),
            "triangularity": _as_float(_safe_get(eq_slice, "boundary.triangularity")),
            "triangularity_upper": _as_float(
                _safe_get(eq_slice, "boundary.triangularity_upper")
            ),
            "triangularity_lower": _as_float(
                _safe_get(eq_slice, "boundary.triangularity_lower")
            ),
            "magnetic_axis_r_m": _as_float(
                _safe_get(eq_slice, "global_quantities.magnetic_axis.r")
            ),
            "magnetic_axis_z_m": _as_float(
                _safe_get(eq_slice, "global_quantities.magnetic_axis.z")
            ),
            "magnetic_axis_btor_T": _as_float(
                _safe_get(eq_slice, "global_quantities.magnetic_axis.b_field_tor")
            ),
            "vacuum_b0_T": normalized_b0,
            "vacuum_r0_m": VACUUM_REFERENCE_RADIUS_M,
            "dia_flux_Wb": _as_float(
                _safe_get(eq_slice, "constraints.diamagnetic_flux.reconstructed")
            ),
            **virial,
        }
        row["virial_beta"] = row["virial_beta_lao"]
        row["virial_li"] = row["virial_li_lao"]
        rows.append(row)
    return rows


def extract_core_profiles(ods, shot: int) -> list[dict]:
    """Extract confinement and engineering parameters on core-profile slices."""
    from vaft.omas import formula_wrapper
    from vaft.omas.general import find_matching_time_indices

    if "equilibrium.time_slice" not in ods or not len(ods["equilibrium.time_slice"]):
        return []
    if "core_profiles.profiles_1d" not in ods or not len(
        ods["core_profiles.profiles_1d"]
    ):
        return []
    try:
        vaft.omas.update_equilibrium_boundary(ods)
    except Exception as exc:
        logger.debug("Shot %s: boundary update failed: %s", shot, exc)

    rows: list[dict] = []
    for cp_index in range(len(ods["core_profiles.profiles_1d"])):
        try:
            matched_cp, eq_index, time_s = find_matching_time_indices(
                ods, time_slice=cp_index
            )
            cp_slice = ods["core_profiles.profiles_1d"][matched_cp]
            engineering = formula_wrapper.compute_tau_E_engineering_parameters(
                ods, eq_index, Z_eff=2.0, M=1.0
            )
            ne_line = float(
                engineering.get("n_e_line_avg", engineering.get("n_e", np.nan))
            )
            ne_volume = float(engineering.get("n_e_vol_avg", np.nan))
            temperature = float(
                np.nanmean(np.asarray(cp_slice["electrons.temperature"], dtype=float))
            )
            if not all(
                np.isfinite(value) and value > 0
                for value in (ne_line, ne_volume, temperature)
            ):
                raise ValueError("density and temperature must be finite and positive")
            (
                tau_ipb89,
                tau_h98y2,
                tau_nstx,
                tau_nstx2006l,
                tau_kurskiev2022,
                _h_factor,
                tau_e,
            ) = formula_wrapper.compute_confiment_time_paramters(
                ods, eq_index, Z_eff=2.0, M=1.0
            )
            rows.append(
                {
                    "shot": int(shot),
                    "cp_index": int(matched_cp),
                    "eq_index": int(eq_index),
                    "time_s": float(time_s),
                    "ip_kA": float(engineering["I_p"]) / 1e3,
                    "b_t_T": float(engineering["B_t"]),
                    "p_loss_MW": float(engineering["P_loss"]) / 1e6,
                    "tau_e_s": float(tau_e),
                    "tau_e_ipb89_s": float(tau_ipb89),
                    "tau_e_h98y2_s": float(tau_h98y2),
                    "tau_e_nstx_s": float(tau_nstx),
                    "tau_e_nstx2006l_s": float(tau_nstx2006l),
                    "tau_e_kurskiev2022_s": float(tau_kurskiev2022),
                    "ne_line_1e19_m3": ne_line / 1e19,
                    "ne_volume_1e19_m3": ne_volume / 1e19,
                    "te_mean_eV": temperature,
                    "major_radius_m": float(engineering["R"]),
                    "inverse_aspect_ratio": float(engineering["epsilon"]),
                    "elongation": float(engineering["kappa"]),
                }
            )
        except Exception as exc:
            logger.warning(
                "Shot %s core_profiles[%s]: extraction failed; skipping: %s",
                shot,
                cp_index,
                exc,
            )
    return rows


def _slice_times(ods, slice_path: str, time_path: str, count: int) -> np.ndarray:
    values = []
    for index in range(count):
        data_slice = ods[slice_path][index]
        if "time" in data_slice:
            values.append(float(data_slice["time"]))
        elif time_path in ods and index < len(ods[time_path]):
            values.append(float(ods[time_path][index]))
        else:
            values.append(float(index))
    return np.asarray(values, dtype=float)


def _pad_values(values, count: int) -> np.ndarray:
    result = np.full(count, np.nan, dtype=float)
    array = np.asarray(values, dtype=float).reshape(-1)
    result[: min(count, array.size)] = array[:count]
    return result


def extract_volume_averaged(ods, shot: int) -> list[dict]:
    """Extract matched core/equilibrium volume-averaged quantities."""
    if "core_profiles.profiles_1d" not in ods or not len(
        ods["core_profiles.profiles_1d"]
    ):
        return []
    if "equilibrium.time_slice" not in ods or not len(ods["equilibrium.time_slice"]):
        return []
    core_count = len(ods["core_profiles.profiles_1d"])
    equilibrium_count = len(ods["equilibrium.time_slice"])
    vaft.omas.update_core_profiles_global_quantities_volume_average(ods)
    global_quantities = ods["core_profiles"].get("global_quantities", {})
    density = _pad_values(global_quantities.get("n_e_volume_average", []), core_count)
    temperature = _pad_values(
        global_quantities.get("t_e_volume_average", []), core_count
    )
    electron_pressure = 2.0 * density * temperature * ELEMENTARY_CHARGE_C
    try:
        equilibrium_pressure = vaft.omas.compute_volume_averaged_pressure(
            ods, time_slice=None, option="equilibrium"
        )
    except Exception as exc:
        logger.warning(
            "Shot %s: equilibrium volume-averaged pressure failed: %s", shot, exc
        )
        equilibrium_pressure = np.full(equilibrium_count, np.nan)
    equilibrium_pressure = _pad_values(equilibrium_pressure, equilibrium_count)
    core_times = _slice_times(
        ods, "core_profiles.profiles_1d", "core_profiles.time", core_count
    )
    equilibrium_times = _slice_times(
        ods, "equilibrium.time_slice", "equilibrium.time", equilibrium_count
    )

    rows: list[dict] = []
    for cp_index, core_time in enumerate(core_times):
        eq_index = int(np.argmin(np.abs(equilibrium_times - core_time)))
        eq_time = float(equilibrium_times[eq_index])
        pressure_core = float(electron_pressure[cp_index])
        pressure_equilibrium = float(equilibrium_pressure[eq_index])
        ratio = (
            pressure_core / pressure_equilibrium
            if np.isfinite(pressure_core)
            and np.isfinite(pressure_equilibrium)
            and pressure_equilibrium != 0
            else np.nan
        )
        rows.append(
            {
                "shot": int(shot),
                "cp_index": int(cp_index),
                "eq_index": int(eq_index),
                "time_core_s": float(core_time),
                "time_equilibrium_s": eq_time,
                "time_diff_s": abs(eq_time - float(core_time)),
                "ne_volume_m3": float(density[cp_index]),
                "te_volume_eV": float(temperature[cp_index]),
                "electron_pressure_volume_Pa": pressure_core,
                "equilibrium_pressure_volume_Pa": pressure_equilibrium,
                "electron_to_equilibrium_pressure_ratio": ratio,
            }
        )
    return rows


def _equilibrium_lineage(ods, eq_index: int) -> str:
    values = {}
    for name, path in (
        ("machine", "dataset_description.data_entry.machine"),
        ("pulse", "dataset_description.data_entry.pulse"),
        ("run", "dataset_description.data_entry.run"),
        ("comment", "equilibrium.ids_properties.comment"),
        ("code_name", "equilibrium.code.name"),
        ("code_version", "equilibrium.code.version"),
        (
            "mapping_source_revision",
            "equilibrium.code.parameters.time_slice."
            f"{eq_index}.meqdsk.mapping_source_revision",
        ),
        (
            "gfile_sha256",
            f"equilibrium.code.parameters.time_slice.{eq_index}.artifacts.gfile.sha256",
        ),
        (
            "mfile_sha256",
            f"equilibrium.code.parameters.time_slice.{eq_index}.artifacts.mfile.sha256",
        ),
    ):
        value = _safe_get(ods, path, None)
        if value is not None:
            try:
                if isinstance(value, np.generic):
                    value = value.item()
                json.dumps(value)
            except (TypeError, ValueError):
                value = str(value)
            values[name] = value
    return json.dumps(values, sort_keys=True, separators=(",", ":"))


def _constraint_indexes(constraint, is_array: bool) -> list[int]:
    if not is_array:
        return [0]
    try:
        if hasattr(constraint, "keys"):
            return sorted(
                int(index) for index in constraint.keys() if str(index).isdigit()
            )
        return list(range(len(constraint)))
    except Exception:
        return []


def _constraint_measurement(
    ods,
    equilibrium_slice,
    eq_index: int,
    family: str,
    measurement_index: int,
    is_array: bool,
):
    constraint = _safe_get(equilibrium_slice, f"constraints.{family}", None)
    if constraint is None:
        return None
    if is_array and hasattr(constraint, "location"):
        return (
            ods,
            f"equilibrium.time_slice.{eq_index}.constraints.{family}."
            f"{measurement_index}",
        )
    try:
        if not is_array:
            return (constraint, "")
        try:
            item = constraint[measurement_index]
        except (KeyError, TypeError, IndexError):
            item = constraint[str(measurement_index)]
        return (item, "")
    except Exception:
        return None


def _measurement_value(measurement, field: str, default=np.nan):
    container, base = measurement
    path = f"{base}.{field}" if base else field
    return _safe_get(container, path, default)


def _extract_efit_reliability_families(
    ods,
    shot: int,
    families: Sequence[tuple[str, bool]],
) -> list[dict]:
    """Extract fitted constraints from finalized equilibrium IDS content."""
    if "equilibrium.time_slice" not in ods or not len(ods["equilibrium.time_slice"]):
        return []
    equilibrium_times = _safe_get(ods, "equilibrium.time", [])
    rows: list[dict] = []
    for eq_index in range(len(ods["equilibrium.time_slice"])):
        equilibrium_slice = ods["equilibrium.time_slice"][eq_index]
        lineage = _equilibrium_lineage(ods, eq_index)
        time_s = _extract_time(equilibrium_slice, equilibrium_times, eq_index)
        ip_kA = _as_float(_safe_get(equilibrium_slice, "global_quantities.ip")) / 1e3
        aggregate_chi_squared = _as_float(
            _safe_get(equilibrium_slice, "constraints.chi_squared_reduced")
        )
        convergence_iterations = _as_float(
            _safe_get(equilibrium_slice, "convergence.iterations_n")
        )
        convergence_deviation = _as_float(
            _safe_get(
                equilibrium_slice,
                "convergence.grad_shafranov_deviation_value",
            )
        )
        for label, is_array in families:
            path = f"constraints.{label}"
            if path not in equilibrium_slice:
                continue
            constraint = equilibrium_slice[path]
            for measurement_index in _constraint_indexes(constraint, is_array):
                measurement = _constraint_measurement(
                    ods,
                    equilibrium_slice,
                    eq_index,
                    label,
                    measurement_index,
                    is_array,
                )
                if measurement is None:
                    continue
                measured = _as_float(_measurement_value(measurement, "measured"))
                reconstructed = _as_float(
                    _measurement_value(measurement, "reconstructed")
                )
                if not (np.isfinite(measured) and np.isfinite(reconstructed)):
                    continue
                uncertainty = _as_float(
                    _measurement_value(measurement, "measured_error_upper")
                )
                weight = _as_float(_measurement_value(measurement, "weight"))
                chi_squared = _as_float(_measurement_value(measurement, "chi_squared"))
                exact_value = _as_float(_measurement_value(measurement, "exact"))
                residual = reconstructed - measured
                normalized = (
                    residual / uncertainty
                    if np.isfinite(uncertainty) and uncertainty != 0.0
                    else np.nan
                )
                identifier = _measurement_value(measurement, "identifier", "")
                if (
                    identifier is None
                    or isinstance(identifier, float)
                    and np.isnan(identifier)
                ):
                    identifier = ""
                source = _measurement_value(measurement, "source", "")
                if source is None or isinstance(source, float) and np.isnan(source):
                    source = ""
                rows.append(
                    {
                        "shot": int(shot),
                        "eq_index": int(eq_index),
                        "time_s": float(time_s),
                        "ip_kA": float(ip_kA),
                        "measurement_index": int(measurement_index),
                        "measurement_type": label,
                        "identifier": str(identifier),
                        "constraint_source": str(source),
                        "measured": measured,
                        "reconstructed": reconstructed,
                        "residual": residual,
                        "absolute_residual": abs(residual),
                        "uncertainty": uncertainty,
                        "normalized_residual": normalized,
                        "weight": weight,
                        "chi_squared": chi_squared,
                        "exact": bool(exact_value)
                        if np.isfinite(exact_value)
                        else np.nan,
                        "disabled": bool(np.isfinite(weight) and weight == 0.0),
                        "chi_squared_reduced": aggregate_chi_squared,
                        "convergence_iterations_n": convergence_iterations,
                        "convergence_deviation": convergence_deviation,
                        "equilibrium_source": "",
                        "equilibrium_lineage": lineage,
                    }
                )
    return rows


def extract_efit_magnetic_reliability(ods, shot: int) -> list[dict]:
    """Extract fitted magnetic constraints from a finalized equilibrium ODS."""
    return _extract_efit_reliability_families(
        ods,
        shot,
        EFIT_MAGNETIC_FAMILIES,
    )


def extract_efit_kinetic_reliability(ods, shot: int) -> list[dict]:
    """Extract fitted kinetic constraints from a finalized equilibrium ODS."""
    return _extract_efit_reliability_families(
        ods,
        shot,
        EFIT_KINETIC_FAMILIES,
    )


def extract_efit_reliability(ods, shot: int) -> list[dict]:
    """Deprecated alias for :func:`extract_efit_magnetic_reliability`."""
    warnings.warn(
        "extract_efit_reliability() is deprecated; use "
        "extract_efit_magnetic_reliability()",
        DeprecationWarning,
        stacklevel=2,
    )
    return extract_efit_magnetic_reliability(ods, shot)


def extract_shot_overview(ods, shot: int) -> list[dict]:
    """Extract one operational overview row from canonical diagnostic signals."""
    from scipy.signal import medfilt

    uv_time = np.asarray(ods["spectrometer_uv.time"], dtype=float)
    uv_intensity = np.asarray(
        ods["spectrometer_uv.channel.0.processed_line.0.intensity.data"],
        dtype=float,
    )
    onset, offset = vaft.process.signal_on_offset(uv_time, uv_intensity, threshold=0.05)
    ip = np.asarray(ods["magnetics.ip.0.data"], dtype=float)
    if not ip.size:
        raise ValueError("magnetics.ip.0.data is empty")
    kernel = min(15, ip.size if ip.size % 2 else ip.size - 1)
    filtered_ip = medfilt(ip, kernel_size=kernel) if kernel >= 3 else ip
    tf_time = np.asarray(ods["tf.time"], dtype=float)
    field = np.asarray(ods["tf.b_field_tor_vacuum_r.data"], dtype=float)
    tf_r0 = _as_float(ods["tf.r0"])
    if field.size != tf_time.size or not np.isfinite(tf_r0) or tf_r0 == 0:
        raise ValueError("invalid tf time, field, or reference radius")
    field_at_reference = field / tf_r0
    in_pulse = (tf_time >= onset) & (tf_time <= offset)
    if not np.any(in_pulse):
        raise ValueError("no toroidal-field samples fall inside the plasma pulse")
    return [
        {
            "shot": int(shot),
            "plasma_onset_time_s": float(onset),
            "pulse_duration_s": float(offset - onset),
            "max_ip_kA": float(np.nanmax(filtered_ip)) / 1e3,
            "mean_b_t_T": float(np.nanmean(field_at_reference[in_pulse])),
        }
    ]


PRESETS = {
    "equilibrium_global": SummaryPreset(
        columns=EQUILIBRIUM_GLOBAL_COLUMNS,
        paths=EQUILIBRIUM_GLOBAL_PATHS,
        key_columns=("shot", "eq_index", "time_s"),
        replace_groups=("shot",),
        sort_columns=("shot", "time_s", "eq_index"),
        extractor=extract_equilibrium_global,
    ),
    "core_profiles": SummaryPreset(
        columns=CORE_PROFILES_COLUMNS,
        paths=CORE_PROFILES_PATHS,
        key_columns=("shot", "cp_index", "time_s"),
        replace_groups=("shot",),
        sort_columns=("shot", "time_s", "cp_index"),
        extractor=extract_core_profiles,
    ),
    "volume_averaged": SummaryPreset(
        columns=VOLUME_AVERAGED_COLUMNS,
        paths=VOLUME_AVERAGED_PATHS,
        key_columns=("shot", "cp_index", "time_core_s"),
        replace_groups=("shot",),
        sort_columns=("shot", "time_core_s", "cp_index"),
        extractor=extract_volume_averaged,
    ),
    "efit_magnetic_reliability": SummaryPreset(
        columns=EFIT_RELIABILITY_COLUMNS,
        paths=("equilibrium",),
        key_columns=("shot", "eq_index", "measurement_type", "measurement_index"),
        replace_groups=("shot",),
        sort_columns=(
            "shot",
            "time_s",
            "eq_index",
            "measurement_type",
            "measurement_index",
        ),
        extractor=extract_efit_magnetic_reliability,
    ),
    "efit_kinetic_reliability": SummaryPreset(
        columns=EFIT_RELIABILITY_COLUMNS,
        paths=("equilibrium",),
        key_columns=("shot", "eq_index", "measurement_type", "measurement_index"),
        replace_groups=("shot",),
        sort_columns=(
            "shot",
            "time_s",
            "eq_index",
            "measurement_type",
            "measurement_index",
        ),
        extractor=extract_efit_kinetic_reliability,
    ),
    "efit_reliability": SummaryPreset(
        columns=EFIT_RELIABILITY_COLUMNS,
        paths=("equilibrium",),
        key_columns=("shot", "eq_index", "measurement_type", "measurement_index"),
        replace_groups=("shot",),
        sort_columns=(
            "shot",
            "time_s",
            "eq_index",
            "measurement_type",
            "measurement_index",
        ),
        extractor=extract_efit_magnetic_reliability,
    ),
    "shot_overview": SummaryPreset(
        columns=SHOT_OVERVIEW_COLUMNS,
        paths=("spectrometer_uv", "magnetics", "tf"),
        key_columns=("shot",),
        replace_groups=("shot",),
        sort_columns=("shot",),
        extractor=extract_shot_overview,
    ),
}


def get_summary_preset(name: str) -> SummaryPreset:
    if name == "efit_reliability":
        warnings.warn(
            "summary preset 'efit_reliability' is deprecated; use "
            "'efit_magnetic_reliability'",
            DeprecationWarning,
            stacklevel=2,
        )
    try:
        return PRESETS[name]
    except KeyError as exc:
        available = ", ".join(sorted(PRESETS))
        raise ValueError(
            f"unknown summary preset {name!r}; available presets: {available}"
        ) from exc


def summary(
    shot_range: tuple[int, int] | None = None,
    *,
    preset: str = "equilibrium_global",
    source: str = "public",
) -> pd.DataFrame:
    """Return a canonical preset summary for a range or every available shot."""
    definition = get_summary_preset(preset)
    from . import _namespace

    source = _namespace(source, "source")
    if shot_range is None:
        from .utils import exist_shot

        discovered = exist_shot(username=source, sort=1) or []
        shots = sorted({int(value) for value in discovered if str(value).isdigit()})
    else:
        if (
            not isinstance(shot_range, (tuple, list))
            or len(shot_range) != 2
            or any(
                isinstance(value, bool) or not isinstance(value, (int, np.integer))
                for value in shot_range
            )
        ):
            raise TypeError("shot_range must be a pair of integer shot numbers")
        start, end = (int(value) for value in shot_range)
        if start > end:
            raise ValueError("shot_range start must be less than or equal to end")
        shots = list(range(start, end + 1))
    frames: list[pd.DataFrame] = []
    from . import open as open_shot

    for shot in shots:
        try:
            with open_shot(shot, source=source, paths=list(definition.paths)) as ods:
                rows = definition.extractor(ods, shot)
            if not rows:
                logger.warning(
                    "Shot %s: preset %s produced no rows; skipping", shot, preset
                )
                continue
            if "equilibrium_source" in definition.columns:
                for row in rows:
                    row["equilibrium_source"] = source
            frames.append(pd.DataFrame(rows))
        except Exception as exc:
            logger.warning("Shot %s: preset %s failed; skipping: %s", shot, preset, exc)

    if not frames:
        return pd.DataFrame(columns=definition.columns)
    result = pd.concat(frames, ignore_index=True)
    for column in definition.columns:
        if column not in result:
            result[column] = np.nan
    return (
        result.loc[:, definition.columns]
        .sort_values(list(definition.sort_columns), kind="stable")
        .reset_index(drop=True)
    )


def _columns(values: Sequence[str] | None, label: str) -> tuple[str, ...]:
    if values is None:
        return ()
    if isinstance(values, str):
        values = (values,)
    result = tuple(values)
    if not result or any(not isinstance(value, str) or not value for value in result):
        raise ValueError(f"{label} must contain non-empty column names")
    return tuple(dict.fromkeys(result))


def _read_table(path: Path) -> pd.DataFrame:
    return pd.read_csv(path) if path.suffix.lower() == ".csv" else pd.read_excel(path)


def _write_table(df: pd.DataFrame, path: Path) -> None:
    if path.suffix.lower() == ".csv":
        df.to_csv(path, index=False)
    else:
        df.to_excel(path, index=False)
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill

        workbook = load_workbook(path)
        sheet = workbook.active
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="17365D")
        for column_index, column in enumerate(df.columns, start=1):
            if pd.api.types.is_numeric_dtype(df[column]):
                is_integer = pd.api.types.is_integer_dtype(df[column])
                number_format = "0" if is_integer else "0.000"
                for row_index in range(2, sheet.max_row + 1):
                    sheet.cell(
                        row=row_index, column=column_index
                    ).number_format = number_format
            header_width = len(str(column)) + 2
            sheet.column_dimensions[
                sheet.cell(1, column_index).column_letter
            ].width = min(max(header_width, 12), 24)
        workbook.save(path)


def export_summary(
    df: pd.DataFrame,
    path: str | os.PathLike,
    *,
    mode: str = "replace",
    key_columns: Sequence[str] | None = None,
    replace_groups: Sequence[str] | None = None,
) -> pd.DataFrame:
    """Serialize a DataFrame to CSV/XLSX, optionally merging an existing table."""
    if not isinstance(df, pd.DataFrame):
        raise TypeError("df must be a pandas DataFrame")
    destination = Path(path)
    if destination.suffix.lower() not in {".csv", ".xlsx"}:
        raise ValueError("export path must end in .csv or .xlsx")
    if mode not in {"replace", "upsert"}:
        raise ValueError("mode must be 'replace' or 'upsert'")

    incoming = df.copy()
    keys = _columns(key_columns, "key_columns")
    groups = _columns(replace_groups, "replace_groups")
    if mode == "replace":
        output = incoming
    else:
        if not keys:
            raise ValueError("key_columns are required for mode='upsert'")
        required = tuple(dict.fromkeys((*keys, *groups)))
        missing = [column for column in required if column not in incoming]
        if missing:
            raise ValueError(f"incoming DataFrame is missing merge columns: {missing}")
        if incoming.loc[:, required].isna().any().any():
            raise ValueError("upsert key and replacement-group values cannot be null")
        if destination.exists():
            existing = _read_table(destination)
            missing_existing = [column for column in required if column not in existing]
            if missing_existing:
                raise ValueError(
                    f"existing export is missing merge columns: {missing_existing}"
                )
            for column in incoming.columns:
                if column not in existing:
                    existing[column] = pd.NA
            existing = existing.loc[:, incoming.columns]
        else:
            existing = pd.DataFrame(columns=incoming.columns)

        if groups and not incoming.empty and not existing.empty:
            incoming_groups = pd.MultiIndex.from_frame(
                incoming.loc[:, groups].drop_duplicates()
            )
            existing_groups = pd.MultiIndex.from_frame(existing.loc[:, groups])
            existing = existing.loc[~existing_groups.isin(incoming_groups)]
        combined = pd.concat([existing, incoming], ignore_index=True)
        output = combined.drop_duplicates(subset=list(keys), keep="last")
        sort_columns = list(dict.fromkeys((*groups, *keys)))
        if not output.empty:
            output = output.sort_values(sort_columns, kind="stable").reset_index(
                drop=True
            )

    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary_name = None
    try:
        with tempfile.NamedTemporaryFile(
            dir=destination.parent,
            prefix=f".{destination.stem}.",
            suffix=destination.suffix,
            delete=False,
        ) as temporary:
            temporary_name = temporary.name
        _write_table(output, Path(temporary_name))
        os.replace(temporary_name, destination)
    finally:
        if temporary_name and os.path.exists(temporary_name):
            os.unlink(temporary_name)
    return output


__all__ = [
    "CORE_PROFILES_COLUMNS",
    "EFIT_RELIABILITY_COLUMNS",
    "EQUILIBRIUM_GLOBAL_COLUMNS",
    "SummaryPreset",
    "export_summary",
    "extract_efit_kinetic_reliability",
    "extract_efit_magnetic_reliability",
    "extract_equilibrium_global",
    "extract_core_profiles",
    "extract_efit_reliability",
    "extract_shot_overview",
    "extract_volume_averaged",
    "get_summary_preset",
    "summary",
    "SHOT_OVERVIEW_COLUMNS",
    "VOLUME_AVERAGED_COLUMNS",
]
